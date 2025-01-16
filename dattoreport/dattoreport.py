import os
import requests
import pyodbc
import json
import base64
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
from openpyxl.styles import NamedStyle

######################################
#        INSTALLED LIBRARIES         #
######################################
# 1. Requests - used for API calls   #
# python -m pip install requests     #
#                                    #
# 2. pyodbc - Used to connect to SQL #
# pip install pyodbc                 #
#                                    #
# 3. dotenv - used to load API data  #
# pip install python-dotenv          #
#                                    #
# 4. openpyxl - used for excel work  #
# pip install openpyxl               #  
######################################
#        ADDITIONAL INSTALLS         #
######################################
# 1. ODBC library                    #
# sudo apt install unixodbc-dev      #
#                                    #
# sudo apt-get update                #
######################################


#############
# dattoinfo #
#############

# loads API key info from .env
# libs - dotenv
load_dotenv()
username = os.getenv('DATTO_PUBLIC_KEY')
password = os.getenv('DATTO_PRIVATE_KEY')

# Build headers to log in and encode
# libs - base64
auth_string = f"{username}:{password}"
encoded_auth_string = base64.b64encode(auth_string.encode()).decode('utf-8')

# url to access and header to use
url = "https://api.datto.com/v1/bcdr" 
headers = {'Authorization': f'Basic {encoded_auth_string}'} 

##############

# Excel - creates date format for excel file
date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD HH:MM:SS")

# Testing - prints json file
# libs - json

def print_json(data):
    try:
        # If the input is a string, attempt to parse it as a json
        if isinstance(data, str):
            data = json.loads(data)

        # Print the JSON object with indents
        print(json.dumps(data, indent=4, sort_keys=True))

    except (json.JSONDecodeError, TypeError) as e:
        print(f"Error: Unable to process the provided data. {e}")

# Datto - call to get list of active devices
# libs - requests, json
def get_active():
    # set url to get device list
    device_url = f"{url}/device"

    try:
        # create list for devices
        devices = [] 
        while device_url:
            # sets response to get device list in response
            response = requests.get(device_url, headers=headers)

            # check return status code
            if response.status_code == 200:
                # parse json response
                data = response.json()
                # print_json(data)
                
                # iterate over clients to get client name
                for item in data.get("items", []):
                    client_name = item.get("organizationName")
                    serial_number = item.get("serialNumber")
                    local_used = item.get("localStorageUsed")
                    local_avail = item.get("localStorageAvailable")

                    if client_name and serial_number:
                        devices.append({
                            "clientName": client_name,
                            "serialNumber": serial_number,
                            "localUsed" : local_used,
                            "localAvail" : local_avail
                        })
                
                # Check for more pages
                device_url = data.get("pagination", {}).get("nextPageUrl")

            
            else:   
                print(f"Error: Failed Device retrieval -  Code:{response.status_code}, Text:{response.text}")
                return None
        
        print(f"Collected devices")
        return devices
    
    except Exception as e:
        print(f"Exception while retrieving devices: {e}")
        return None
    
# Datto - call to get device backup info
# libs - requests, json
def get_backups(serialNumber):
    # set url to get device info
    backup_url = f"{url}/device/{serialNumber}/asset"

    try:
        # sets response to get device status
        response = requests.get(backup_url, headers=headers)

        # Check return code
        if response.status_code == 200:
            # if success - return json
            return response.json()
        
        else:
            print(f"Error: Could not retrieve info for {serialNumber} Error {response.status_code}, {response.text}")
            return None
        
    except Exception as e:
        print(f"Exception during retrieval for {serialNumber}: {e}")
        return None
    
# Excel - used to convert date strings into readable formats
# libs - datetime
def parse_date(date_str):
    # check for valid args
    if not date_str:
        return None
    
    # converts time in seconds/units to UTC *Note - has to be timezone naive for excel file
    try:
        # gets raw time counter
        timestamp = int(date_str)
        # converts time counter to a date time format
        dt = datetime.fromtimestamp(timestamp, timezone.utc)
        # strips timezone and returns to use
        return dt.replace(tzinfo=None)
    
    except (ValueError, TypeError):
        pass

    # converts time in string format parsed as ISO 8601
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    # returns None if no values found
    return None

# Excel - used to simplify data used expressions to single string    
def parse_storage(storage): 
    # type checking for error handling
    if not isinstance(storage, dict):
        return "Cant Convert Storage Total - Not a Dictionary"

    size = storage.get('size', 'Unknown')
    units = storage.get('units', 'Unknown')        
    return f"{size} {units}"
    
# Excel - writes data to excel workbook using pre-defined headers
# libs -  openpyxl, datetime
def write_xlsx(device_backup_data, filename="datto_report.xlsx"):
    # Looks for template defined in load_workbook
    print(f"Loading Template & Writing Report")
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        print(f"Template Not Found")

    # Check for second page of workbook
    if len(wb.sheetnames) < 2:
        print(f"Template missing second page")
        return

    # sets second page as active sheet
    ws = wb[wb.sheetnames[1]]
    output_filename = None

    # define headers for columns
    headers = [
        "serialNumber", "Client Name", "Datto Name", "Server Name", "agentVersion", "isPaused", "isArchived", 
        "latestOffsite", "lastSnapshot", "lastScreenshotAttempt", 
        "lastScreenshotAttemptStatus", "lastScreenshotUrl", 
        "localStorageUsed", "localStorageAvailable"
    ]

    # Write headers to first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Write each device's backup data to following rows
    for row_num, data in enumerate(device_backup_data, 2):  # Start from row 2
        ws[f"A{row_num}"] = data["serialNumber"]
        ws[f"B{row_num}"] = data["Client Name"]
        ws[f"C{row_num}"] = data["Datto Name"]
        ws[f"D{row_num}"] = data["Server Name"]
        ws[f"E{row_num}"] = data["agentVersion"]
        ws[f"F{row_num}"] = data["isPaused"]
        ws[f"G{row_num}"] = data["isArchived"]

        # apply date style for relevant fields
        for date_field in ["latestOffsite", "lastSnapshot", "lastScreenshotAttempt"]:
            if data.get(date_field):
                cell = ws[f"{get_column_letter(headers.index(date_field)+1)}{row_num}"]
                cell.value = data[date_field]
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

        ws[f"K{row_num}"] = data["lastScreenshotAttemptStatus"]
        ws[f"L{row_num}"] = data["lastScreenshotUrl"]
        ws[f"M{row_num}"] = data["localStorageUsed"]
        ws[f"N{row_num}"] = data["localStorageAvailable"]

    # Save workbook
    if output_filename is None:
        current_date = datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD
        output_filename = f"datto_report_{current_date}.xlsx"

    wb.save(output_filename)
    print(f"Data written to {output_filename}")


# Datto - API call process function
# Reaches out to datto to get list of devices
# Iterates over devices to get serial numbers
# Returns dict with nested dicts full of client backup info
def datto_report():
    # get list of active devices
    print(f"Retrieving Device list")
    devices = get_active()

    # Check for return of None to report error
    if devices is None or not devices:
        print("No devices found/retrieved")
        return

    full_backup_data = []
    print(f"Polling devices for backup data")
    
    # Iterate over devices
    for device in devices:
        # Get serialNumber for device
        serialNumber = device.get("serialNumber")
        client_name = device.get("clientName", "")
        device_name = device.get("name", "")
        local_used = parse_storage(device.get("localUsed", ""))
        local_avail = parse_storage(device.get("localAvail", ""))
        # if serial is found check for backups
        if serialNumber:
            backup_info = get_backups(serialNumber)          

            # if backups are found
            if backup_info:
                # check to see if info is in proper format
                if isinstance(backup_info, list):
                    for backup in backup_info:
                        device_data = {
                            "serialNumber": serialNumber,
                            "Client Name": client_name,
                            "Datto Name" : device_name,
                            "Server Name": backup.get("name", ""),
                            "agentVersion": backup.get("agentVersion", ""),
                            "isPaused": backup.get("isPaused", ""),
                            "isArchived": backup.get("isArchived", ""),
                            "latestOffsite": parse_date(backup.get("latestOffsite", "")),
                            "lastSnapshot": parse_date(backup.get("lastSnapshot", "")),
                            "lastScreenshotAttempt": parse_date(backup.get("lastScreenshotAttempt", "")),
                            "lastScreenshotAttemptStatus": backup.get("lastScreenshotAttemptStatus", ""),
                            "lastScreenshotUrl": backup.get("lastScreenshotUrl", ""),
                            "localStorageUsed": local_used,
                            "localStorageAvailable": local_avail,
                        }
                        
                        # Append the device's data to the list
                        full_backup_data.append(device_data)
                # error check for backup data type (looking for list)
                else:
                    print(f"Unexpected format for backup data: {type(backup_info)}")
            # error check for if backup data is present for device at serialNumber
            else: 
                print(f"No backup info found for {serialNumber}")
        # error check for presence of serialNumber
        else:
            print(f"Device serial not found")
            
    return full_backup_data


if __name__ == "__main__":
    print("Running report directly")
    write_xlsx(datto_report())
            